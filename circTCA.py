import tensorflow as tf
from d2l import tensorflow as d2l
from sklearn.model_selection import KFold
import argparse
from keras import layers
from keras.models import Model
from keras.layers import Input, Embedding
from keras.layers import Convolution1D, Concatenate, Bidirectional, LSTM, GRU, Flatten, AvgPool1D,\
    BatchNormalization, Activation, Dense, ReLU, MaxPooling1D
from tcn import TCN
import openpyxl
import pandas
from keras.callbacks import EarlyStopping,ReduceLROnPlateau
import numpy as np
from one_hot import *
from multiattention import *
from multiattention2 import *
from ePooling import *
from attention import *

def createModel():
    sequence_input1 = Input(shape=(101, 4), name='sequence_input1')
    sequence_input2 = Input(shape=(101, 64), name='sequence_input2')

    #TCN
    block1 = TCN(nb_filters=128, kernel_size=10, dropout_rate=0.3, nb_stacks=1,
                dilations=[1,2,4,8], return_sequences=True,activation='relu',padding='same',
                use_skip_connections=True)(sequence_input1)
    block1 = MaxPooling1D(pool_size=5)(block1)
    #TCN
    block2 = TCN(nb_filters=128, kernel_size=10, dropout_rate=0.3, nb_stacks=1,
                dilations=[1,2,4,8], return_sequences=True,activation='relu',padding='same',
                use_skip_connections=True)(sequence_input2)
    block2 = MaxPooling1D(pool_size=5)(block2)

    attention = MultiHeadAttention()
    block3 = attention(block1, block2)
    block4 = attention(block2, block1)

    block5 = Concatenate(axis=-1)([block3, block4])
    block6 = GlobalExpectationPooling1D(mode = 0, m_trainable = False,
                                              m_value = 1)(block5)
    output = Dense(2, activation='sigmoid', name='output')(block6)
    return Model(inputs=[sequence_input1, sequence_input2], outputs=[output])

def train_model(protein, parser, trainX1, trainX2, testX1, testX2, trainy, testy):
    protein = protein
    batch_size = parser.batch_size
    n_epochs = parser.n_epochs
    trainXeval1 = trainX1
    trainXeval2 = trainX2
    trainYeval = trainy
    test_X1 = testX1
    test_X2 = testX2
    test_y = testy[:, 1]

    kf = KFold(n_splits=5)
    id = 0
    for train_index, eval_index in kf.split(trainYeval):
        id = id+1
        train_X1 = trainXeval1[train_index]
        train_X2 = trainXeval2[train_index]
        train_y = trainYeval[train_index]
        eval_X1 = trainXeval1[eval_index]
        eval_X2 = trainXeval2[eval_index]
        eval_y = trainYeval[eval_index]
        print('configure network')
        model = createModel()
        model.compile(optimizer='Adam',
                      loss={'output': 'binary_crossentropy'},
                      metrics='binary_accuracy')
        model.summary()
        earlystopper = EarlyStopping(monitor='val_loss', patience=5, verbose=0)
        reduce_lr = ReduceLROnPlateau(monitor='val_accuracy', patience=5, mode='auto')  # Automatically adjust the learning rate
        model.fit({'sequence_input1': train_X1, 'sequence_input2': train_X2}, {'output': train_y},
            epochs= n_epochs,
            batch_size= batch_size,
            callbacks= [earlystopper],
            verbose = 1,
            validation_data=({'sequence_input1': eval_X1, 'sequence_input2': eval_X2}, {'output': eval_y}))
        predictions = model.predict({'sequence_input1': test_X1, 'sequence_input2': test_X2})
        ypred = predictions[:, 1]

        #store results
        f = openpyxl.Workbook()
        sheet = f.active
        # 生成字段名（第一行）
        count = 0
        for i in range(len(ypred)):
            sheet.cell(row=count+1, column=0+1, value=str(test_y[count]))
            sheet.cell(row=count+1, column=0+2, value=str(ypred[count]))
            count += 1
        f.save(protein + str(id) + '.xlsx')
        data = pandas.read_excel(protein + str(id)+ '.xlsx', index_col=0)
        data.to_csv(protein + str(id) +'.csv', encoding='utf-8')

def parse_arguments(parser):
    parser.add_argument('--protein', type=str, default='WTAP')
    parser.add_argument('--batch_size', type=int, default=50)
    parser.add_argument('--n_epochs', type=int, default=500)
    args = parser.parse_args()
    return args

def getprotein(parser):
    protein = parser.protein
    return protein

if __name__ == '__main__':
     parser = argparse.ArgumentParser()
     args = parse_arguments(parser)
     protein = getprotein(args)
     trainXeval1, trainXeval2, test_X1, test_X2, trainYeval, test_y = dealwithdata(protein)
     train_model(protein, args, trainXeval1, trainXeval2, test_X1, test_X2, trainYeval,test_y)





